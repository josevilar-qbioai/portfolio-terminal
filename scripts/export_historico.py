#!/usr/bin/env python3
"""
export_historico.py — Exporta el histórico de precios a Excel y Markdown
══════════════════════════════════════════════════════════════════════════
Uso:
  python3 export_historico.py                        # Exporta todo
  python3 export_historico.py -f portfolio.yaml      # YAML concreto
  python3 export_historico.py --desde 2024-01-01     # Rango de fechas
  python3 export_historico.py --solo-md              # Solo informe MD
  python3 export_historico.py --solo-xlsx            # Solo Excel

Genera:
  exports/historico_portfolio_YYYY-MM-DD.xlsx   ← Excel con gráficos
  exports/historico_portfolio_YYYY-MM-DD.md     ← Informe Markdown
"""

import csv
import yaml
import argparse
import sys
from datetime import date, datetime
from pathlib import Path

# ── CSV reader ─────────────────────────────────────────────────────────────────
def read_csv(path: Path, desde: str = None, hasta: str = None) -> list[dict]:
    if not path.exists():
        return []
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            fecha = row.get("fecha", "")
            if desde and fecha < desde:
                continue
            if hasta and fecha > hasta:
                continue
            rows.append(row)
    return sorted(rows, key=lambda r: r["fecha"])

def rentabilidad(rows: list[dict], desde: str = None) -> dict:
    """Calcula métricas de rentabilidad sobre la serie de precios."""
    precios = [(r["fecha"], float(r["precio_cierre"])) for r in rows if r.get("precio_cierre")]
    if not precios:
        return {}

    p_first_date, p_first = precios[0]
    p_last_date,  p_last  = precios[-1]
    p_min = min(p for _, p in precios)
    p_max = max(p for _, p in precios)
    total_pct = (p_last - p_first) / p_first * 100 if p_first else 0

    # Rentabilidad en distintos períodos
    hoy = date.today().isoformat()
    periodos = {}
    for label, dias in [("1M", 30), ("3M", 90), ("6M", 180), ("1A", 365), ("2A", 730)]:
        import datetime as dt
        desde_fecha = (dt.date.today() - dt.timedelta(days=dias)).isoformat()
        subset = [(d, p) for d, p in precios if d >= desde_fecha]
        if len(subset) >= 2:
            p0, p1 = subset[0][1], subset[-1][1]
            periodos[label] = (p1 - p0) / p0 * 100 if p0 else None
        else:
            periodos[label] = None

    return {
        "n":            len(precios),
        "fecha_inicio": p_first_date,
        "fecha_fin":    p_last_date,
        "precio_inicio":p_first,
        "precio_actual":p_last,
        "precio_min":   p_min,
        "precio_max":   p_max,
        "total_pct":    total_pct,
        "periodos":     periodos,
        "series":       precios,
    }

# ── Excel export ───────────────────────────────────────────────────────────────
def export_xlsx(assets_data: dict, output_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Color palette ──────────────────────────────────────────────────────────
    BG_DARK   = "0D0D0D"
    BG_HEADER = "1A1A1A"
    BG_ROW1   = "0F0F0F"
    BG_ROW2   = "141414"
    AMBER     = "FF9900"
    GREEN     = "00CC66"
    RED       = "FF3333"
    CYAN      = "00CCFF"
    WHITE     = "E8E0CC"
    DIM       = "555555"

    thin = Side(style="thin", color="222222")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_style(cell, text, color=AMBER):
        cell.value = text
        cell.font = Font(name="Consolas", bold=True, color=color, size=9)
        cell.fill = PatternFill("solid", fgColor=BG_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def data_style(cell, value, color=WHITE, bold=False, num_fmt=None, align="right"):
        cell.value = value
        cell.font = Font(name="Consolas", color=color, bold=bold, size=9)
        cell.fill = PatternFill("solid", fgColor=BG_ROW1)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border
        if num_fmt:
            cell.number_format = num_fmt

    def pct_color(v):
        if v is None: return DIM
        return GREEN if v >= 0 else RED

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1: RESUMEN
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("RESUMEN")
    ws.sheet_view.showGridLines = False
    ws.tab_color = AMBER

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"].value = "PORTAFOLIO · HISTÓRICO DE PRECIOS"
    ws["A1"].font = Font(name="Consolas", bold=True, color=AMBER, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="080808")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"Generado: {date.today().isoformat()}"
    ws["A2"].font = Font(name="Consolas", color=DIM, size=8)
    ws["A2"].fill = PatternFill("solid", fgColor="080808")

    # Headers row
    headers = ["TICKER", "DESDE", "HASTA", "Nº DÍAS", "PRECIO INICIO", "PRECIO ACTUAL",
               "MÍN", "MÁX", "TOTAL %"]
    for col, h in enumerate(headers, 1):
        hdr_style(ws.cell(4, col), h)
    ws.row_dimensions[4].height = 22

    # Data rows
    row = 5
    for ticker, info in assets_data.items():
        r = info["rentab"]
        if not r:
            continue
        cells_data = [
            (ticker,               CYAN,  False, None, "left"),
            (r["fecha_inicio"],    WHITE, False, None, "center"),
            (r["fecha_fin"],       WHITE, False, None, "center"),
            (r["n"],               WHITE, False, None, "right"),
            (r["precio_inicio"],   WHITE, False, '#,##0.0000', "right"),
            (r["precio_actual"],   WHITE, True,  '#,##0.0000', "right"),
            (r["precio_min"],      DIM,   False, '#,##0.0000', "right"),
            (r["precio_max"],      DIM,   False, '#,##0.0000', "right"),
            (r["total_pct"]/100,   pct_color(r["total_pct"]), True, '+0.00%;-0.00%;-', "right"),
        ]
        for col, (val, color, bold, fmt, align) in enumerate(cells_data, 1):
            data_style(ws.cell(row, col), val, color, bold, fmt, align)
        ws.row_dimensions[row].height = 18
        row += 1

    # Period performance table
    row += 1
    ws.merge_cells(f"A{row}:I{row}")
    ws.cell(row, 1).value = "RENTABILIDAD POR PERÍODO"
    ws.cell(row, 1).font = Font(name="Consolas", bold=True, color=AMBER, size=10)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=BG_HEADER)
    ws.row_dimensions[row].height = 20
    row += 1

    period_hdrs = ["TICKER"] + ["1M", "3M", "6M", "1A", "2A"]
    for col, h in enumerate(period_hdrs, 1):
        hdr_style(ws.cell(row, col), h)
    row += 1

    for ticker, info in assets_data.items():
        r = info["rentab"]
        if not r:
            continue
        data_style(ws.cell(row, 1), ticker, CYAN, False, None, "left")
        for col, period in enumerate(["1M", "3M", "6M", "1A", "2A"], 2):
            v = r["periodos"].get(period)
            if v is not None:
                data_style(ws.cell(row, col), v/100, pct_color(v), False, '+0.00%;-0.00%;-')
            else:
                data_style(ws.cell(row, col), "—", DIM, False, None, "center")
        ws.row_dimensions[row].height = 18
        row += 1

    # Column widths
    widths = [12, 12, 12, 8, 14, 14, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA por activo: datos + gráfico
    # ══════════════════════════════════════════════════════════════════════════
    chart_colors = [AMBER, CYAN, "00FF80", "FF66CC", "FFCC00", "FF6633"]

    for idx, (ticker, info) in enumerate(assets_data.items()):
        rows_data = info["rows"]
        if not rows_data:
            continue

        safe_name = ticker.replace(".", "_").replace("/", "-")[:28]
        ws2 = wb.create_sheet(safe_name)
        ws2.sheet_view.showGridLines = False
        ws2.tab_color = chart_colors[idx % len(chart_colors)]

        # Title
        ws2.merge_cells("A1:E1")
        ws2["A1"].value = ticker
        ws2["A1"].font = Font(name="Consolas", bold=True, color=chart_colors[idx % len(chart_colors)], size=13)
        ws2["A1"].fill = PatternFill("solid", fgColor="080808")
        ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[1].height = 24

        # Subtitle
        r = info["rentab"]
        if r:
            ws2.merge_cells("A2:E2")
            ws2["A2"].value = f"{r['n']} registros · {r['fecha_inicio']} → {r['fecha_fin']} · Total: {r['total_pct']:+.2f}%"
            ws2["A2"].font = Font(name="Consolas", color=DIM, size=8)
            ws2["A2"].fill = PatternFill("solid", fgColor="080808")

        # Headers
        for col, h in enumerate(["FECHA", "PRECIO CIERRE", "VARIACIÓN DÍA", "VAR. %", "FUENTE"], 1):
            hdr_style(ws2.cell(4, col), h)
        ws2.row_dimensions[4].height = 20

        # Data
        prev_price = None
        for i, row_d in enumerate(rows_data, 5):
            precio = float(row_d["precio_cierre"])
            var_abs = precio - prev_price if prev_price is not None else 0
            var_pct = var_abs / prev_price if prev_price else 0
            var_color = GREEN if var_abs >= 0 else RED

            data_style(ws2.cell(i, 1), row_d["fecha"], DIM, False, None, "center")
            data_style(ws2.cell(i, 2), precio, WHITE, True, '#,##0.0000')
            if prev_price is not None:
                data_style(ws2.cell(i, 3), var_abs, var_color, False, '+#,##0.0000;-#,##0.0000;-')
                data_style(ws2.cell(i, 4), var_pct, var_color, False, '+0.00%;-0.00%;-')
            else:
                data_style(ws2.cell(i, 3), "—", DIM, False, None, "center")
                data_style(ws2.cell(i, 4), "—", DIM, False, None, "center")
            data_style(ws2.cell(i, 5), row_d.get("fuente", ""), DIM, False, None, "center")
            ws2.row_dimensions[i].height = 16
            prev_price = precio

        # ── Line chart ────────────────────────────────────────────────────────
        n_rows = len(rows_data)
        if n_rows >= 2:
            chart = LineChart()
            chart.title    = f"{ticker} · Evolución precio"
            chart.style    = 10
            chart.y_axis.title = "Precio"
            chart.x_axis.title = "Fecha"
            chart.height   = 12
            chart.width    = 22
            chart.grouping = "standard"

            data_ref = Reference(ws2, min_col=2, max_col=2, min_row=4, max_row=4 + n_rows)
            chart.add_data(data_ref, titles_from_data=True)

            cats = Reference(ws2, min_col=1, min_row=5, max_row=4 + n_rows)
            chart.set_categories(cats)

            s = chart.series[0]
            s.graphicalProperties.line.solidFill = chart_colors[idx % len(chart_colors)]
            s.graphicalProperties.line.width = 15000  # 1.5pt

            ws2.add_chart(chart, "G4")

        # Column widths
        for col, w in enumerate([13, 14, 14, 10, 10], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    print(f"  ✓ Excel guardado: {output_path}")

# ── Markdown export ────────────────────────────────────────────────────────────
def export_md(assets_data: dict, portfolio_name: str, output_path: Path):
    today = date.today().isoformat()
    lines = []

    lines += [
        f"# 📈 {portfolio_name} · Histórico de Precios",
        f"> Generado: {today}",
        "",
        "---",
        "",
        "## Resumen de Rentabilidades",
        "",
        "| Activo | Desde | Registros | Precio inicio | Precio actual | Total | 1M | 3M | 1A |",
        "|--------|-------|-----------|--------------|---------------|-------|----|----|----|",
    ]

    def pct_fmt(v):
        if v is None: return "—"
        sign = "+" if v >= 0 else ""
        return f"{sign}{v:.2f}%"

    for ticker, info in assets_data.items():
        r = info["rentab"]
        if not r:
            continue
        lines.append(
            f"| **{ticker}** | {r['fecha_inicio']} | {r['n']} | "
            f"{r['precio_inicio']:.4f} | **{r['precio_actual']:.4f}** | "
            f"**{pct_fmt(r['total_pct'])}** | "
            f"{pct_fmt(r['periodos'].get('1M'))} | "
            f"{pct_fmt(r['periodos'].get('3M'))} | "
            f"{pct_fmt(r['periodos'].get('1A'))} |"
        )

    lines += ["", "---", "", "## Detalle por Activo", ""]

    for ticker, info in assets_data.items():
        rows = info["rows"]
        r    = info["rentab"]
        if not rows or not r:
            continue

        lines += [
            f"### {ticker}",
            "",
            f"- **Rango:** {r['fecha_inicio']} → {r['fecha_fin']}",
            f"- **Registros:** {r['n']} días",
            f"- **Precio mín/máx:** {r['precio_min']:.4f} / {r['precio_max']:.4f}",
            f"- **Variación total:** {pct_fmt(r['total_pct'])}",
            "",
        ]

        # Sparkline
        precios = [float(row["precio_cierre"]) for row in rows]
        if len(precios) >= 2:
            mn, mx = min(precios), max(precios)
            rng = mx - mn or 1
            bars = "▁▂▃▄▅▆▇█"
            spark = "".join(bars[min(7, int((p - mn) / rng * 7))] for p in precios[-40:])
            lines += [f"```", f"Tendencia: {spark}", f"```", ""]

        # Rentabilidad por periodo
        lines += ["| Período | Rentabilidad |", "|---------|-------------|"]
        for label in ["1M", "3M", "6M", "1A", "2A"]:
            v = r["periodos"].get(label)
            lines.append(f"| {label} | {pct_fmt(v)} |")
        lines += [""]

        # Last 10 prices table
        last10 = rows[-10:]
        lines += [
            "#### Últimas 10 sesiones",
            "",
            "| Fecha | Precio | Var. día |",
            "|-------|--------|---------|",
        ]
        prev = None
        for row in last10:
            p = float(row["precio_cierre"])
            if prev is not None:
                var = (p - prev) / prev * 100
                var_str = pct_fmt(var)
            else:
                var_str = "—"
            lines.append(f"| {row['fecha']} | {p:.4f} | {var_str} |")
            prev = p

        lines += ["", "---", ""]

    lines += [
        "## Notas",
        "",
        "- Precios grabados diariamente al cierre de mercado.",
        "- Fuente: `yfinance` (tiempo real) o `precios_actuales` del YAML (manual).",
        f"- Exportado el {today} por `export_historico.py`.",
        "",
        f"*{portfolio_name} · {today}*",
    ]

    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  ✓ Markdown guardado: {output_path}")

# ── Entry point ────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Portfolio History Exporter")
    parser.add_argument("-f",  "--file",       default="portfolio.yaml")
    parser.add_argument("-d",  "--dir",        default="historico",      help="Directorio con CSVs")
    parser.add_argument("-o",  "--output-dir", default="exports",        help="Directorio de salida")
    parser.add_argument("--desde",             default=None,             help="Fecha inicio YYYY-MM-DD")
    parser.add_argument("--hasta",             default=None,             help="Fecha fin YYYY-MM-DD")
    parser.add_argument("--solo-md",           action="store_true")
    parser.add_argument("--solo-xlsx",         action="store_true")
    args = parser.parse_args()

    yaml_path  = Path(args.file)
    hist_dir   = Path(args.dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not yaml_path.exists():
        print(f"\n  ✗ No se encuentra: {yaml_path}\n")
        sys.exit(1)

    with open(yaml_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    activos        = data.get("activos") or []
    portfolio_name = data.get("meta", {}).get("nombre", "Mi Portafolio")
    today          = date.today().isoformat()

    print(f"\n{'═'*55}")
    print(f"  PORTFOLIO HISTORY EXPORTER · {today}")
    print(f"  Portafolio: {portfolio_name}")
    print(f"{'═'*55}\n")

    # Load all asset data
    assets_data = {}
    for a in activos:
        ticker   = a["ticker"]
        csv_path = hist_dir / f"{ticker.replace('/', '-')}.csv"
        rows     = read_csv(csv_path, args.desde, args.hasta)
        assets_data[ticker] = {
            "rows":   rows,
            "rentab": rentabilidad(rows, args.desde),
        }
        n = len(rows)
        print(f"  ← {ticker:<12} {n} registros {'(sin datos)' if n == 0 else ''}")

    print()
    stamp = today

    if not args.solo_md:
        xlsx_path = output_dir / f"historico_portfolio_{stamp}.xlsx"
        export_xlsx(assets_data, xlsx_path)

    if not args.solo_xlsx:
        md_path = output_dir / f"historico_portfolio_{stamp}.md"
        export_md(assets_data, portfolio_name, md_path)

    print(f"\n  ✓ Exportación completada → {output_dir}/")
    print(f"{'═'*55}\n")

if __name__ == "__main__":
    main()
